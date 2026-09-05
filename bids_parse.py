"""Parse IT DIR / Financial Year Bids Excel workbooks into bid line packs."""
from __future__ import annotations

import datetime
import io
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


# Sheet-name hints → system GL (legacy consumables code kept for bids packs)
SHEET_GL_HINTS = [
    (re.compile(r"toner|consumab|zoff|office", re.I), "2200600002"),
    (re.compile(r"software|licence|license", re.I), "2200600003"),
    (re.compile(r"parts|spare", re.I), "2201900002"),
    (re.compile(r"maint|tech\s*eqpt|maintenance", re.I), "220200002"),
    (re.compile(r"\bict\b|equipment|laptop|desktop|server|printer", re.I), "3112210001"),
]

SKIP_SHEET = re.compile(r"summary|cover|index|notes|readme|instructions|chart|pivot", re.I)

KNOWN_GLS = {
    "6122100009",
    "2200600002",
    "2200600003",
    "220200002",
    "2202000002",
    "2202000004",
    "2201900002",
    "3112210001",
}

GL_ALIASES = {
    "2202000002": "220200002",
    "2202000004": "220200002",
    "2200600002": "2200600002",  # consumables (legacy in bids)
}


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("USD", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        return float(m.group(0)) if m else 0.0


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def normalize_gl(raw: str, sheet_name: str = "") -> str:
    digits = re.sub(r"\D", "", str(raw or ""))
    if digits in KNOWN_GLS or len(digits) >= 7:
        gl = digits if digits in KNOWN_GLS else digits
        return GL_ALIASES.get(gl, gl)
    for rx, gl in SHEET_GL_HINTS:
        if rx.search(sheet_name or ""):
            return gl
    return "3112210001"


def infer_fy(filename: str, sheet_name: str = "") -> str:
    text = f"{filename} {sheet_name}"
    m = re.search(r"20[2-3]\d", text)
    if m:
        return m.group(0)
    return str(datetime.date.today().year)


def header_key(text: str) -> str:
    t = re.sub(r"\s+", " ", str(text or "").strip().lower())
    t = t.replace("/", " ").replace("-", " ")
    return t


def classify_header(text: str) -> str | None:
    h = header_key(text)
    if not h:
        return None
    if h in {"ser", "serial", "s/n", "sn", "#", "no", "no."}:
        return "serial"
    if "cost centre" in h or "cost center" in h or h in {"cc", "c/c", "costcentre"}:
        return "costCentre"
    if h.startswith("gl") or "gl a" in h or "g/l" in h or "vote" in h or "account" == h:
        return "gl"
    if "unit cost" in h or "unit price" in h or h in {"price", "rate", "unitcost"}:
        return "unitCost"
    if "total cost" in h or h in {"total", "amount", "line total", "extended"}:
        return "totalCost"
    if h in {"qty", "quantity", "qnty", "quantity required", "qty required", "no required"}:
        return "qty"
    if "justif" in h or "description" in h or "remarks" in h or "purpose" in h:
        return "description"
    if h in {"item", "items", "item name", "nomenclature", "particulars", "stores"} or h.startswith("item "):
        return "item"
    return None


def score_header_row(cells: list) -> int:
    mapped = [classify_header(c) for c in cells]
    keys = {m for m in mapped if m}
    score = 0
    if "item" in keys:
        score += 3
    if "qty" in keys:
        score += 2
    if "unitCost" in keys:
        score += 2
    if "totalCost" in keys:
        score += 1
    if "gl" in keys:
        score += 1
    if "description" in keys:
        score += 1
    return score


def find_header(rows: list[list], max_scan: int = 25) -> tuple[int, dict[str, int]] | None:
    best = None
    for idx, row in enumerate(rows[:max_scan]):
        cells = [cell_str(c) for c in row]
        score = score_header_row(cells)
        if score < 4:
            continue
        colmap: dict[str, int] = {}
        for i, cell in enumerate(cells):
            key = classify_header(cell)
            if key and key not in colmap:
                colmap[key] = i
        if "item" not in colmap and "description" in colmap:
            colmap["item"] = colmap["description"]
        if "item" not in colmap:
            continue
        cand = (score, idx, colmap)
        if best is None or cand[0] > best[0]:
            best = cand
    if not best:
        return None
    return best[1], best[2]


def is_junk_item(name: str) -> bool:
    n = (name or "").strip().lower()
    if not n:
        return True
    if n in {"total", "sub total", "subtotal", "grand total", "totals", "nil", "n/a", "-"}:
        return True
    if n.startswith("total ") or "grand total" in n:
        return True
    return False


def row_to_item(cells: list, colmap: dict[str, int], *, sheet: str, fy: str, source: str) -> dict | None:
    def get(key: str, default=""):
        i = colmap.get(key)
        if i is None or i >= len(cells):
            return default
        return cells[i]

    item = cell_str(get("item"))
    if is_junk_item(item):
        return None

    desc = cell_str(get("description"))
    if desc and desc.lower() == item.lower():
        desc = ""

    qty = num(get("qty"))
    unit = num(get("unitCost"))
    total = num(get("totalCost"))
    if total <= 0 and qty and unit:
        total = round(qty * unit, 2)
    if unit <= 0 and qty and total:
        unit = round(total / qty, 4)
    # Need at least a name; allow zero qty for planning lines still listed
    if not item:
        return None
    # Skip empty numeric rows that look like section headers with no figures
    if qty <= 0 and unit <= 0 and total <= 0 and not desc and len(item) < 3:
        return None

    cc = cell_str(get("costCentre")) or "Z04P2SP212"
    gl = normalize_gl(cell_str(get("gl")), sheet)

    return {
        "item": item,
        "description": desc,
        "costCentre": cc,
        "gl": gl,
        "qty": qty,
        "unitCost": unit,
        "totalCost": total,
        "sheet": sheet,
        "fy": fy,
        "source": source,
    }


def parse_sheet(ws, *, sheet_name: str, fy: str, source: str) -> list[dict]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    found = find_header(rows)
    if not found:
        # Fallback: assume first non-empty row is header-ish Ser | Item | Qty | Unit | Total
        # Try common fixed layout starting at row 1–5
        for start in range(min(8, len(rows))):
            probe = rows[start]
            non_empty = sum(1 for c in probe if cell_str(c))
            if non_empty >= 3:
                # treat as data without header if second cell looks like an item
                pass
        return []

    header_idx, colmap = found
    items: list[dict] = []
    for row in rows[header_idx + 1 :]:
        if not row or not any(row):
            continue
        cells = list(row)
        rec = row_to_item(cells, colmap, sheet=sheet_name, fy=fy, source=source)
        if rec:
            items.append(rec)
    return items


def pack_id_from_name(filename: str) -> str:
    base = Path(filename).stem
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower()
    return slug[:64] or f"import-{datetime.date.today().isoformat()}"


def parse_bids_workbook(wb, *, source_name: str = "") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    source = source_name or "bids.xlsx"
    fy = infer_fy(source)
    items: list[dict] = []
    sheets_used: list[str] = []
    sheets_skipped: list[str] = []

    for name in wb.sheetnames:
        if SKIP_SHEET.search(name or ""):
            sheets_skipped.append(name)
            continue
        ws = wb[name]
        sheet_items = parse_sheet(ws, sheet_name=name, fy=infer_fy(source, name), source=Path(source).stem)
        if not sheet_items:
            sheets_skipped.append(name)
            continue
        sheets_used.append(name)
        items.extend(sheet_items)

    if not items:
        raise ValueError(
            "No bid lines found. Expect sheets with columns such as Item, Quantity, Unit Cost / Total "
            "(IT DIR BIDS layout), or a flat table matching Financial Year Bids."
        )

    by_gl: dict[str, float] = {}
    for it in items:
        by_gl[it["gl"]] = by_gl.get(it["gl"], 0.0) + float(it.get("totalCost") or 0)

    label = Path(source).stem.replace("_", " ").strip() or "Imported bids"
    pid = pack_id_from_name(source)

    return {
        "id": pid,
        "label": label,
        "fy": fy,
        "file": Path(source).name,
        "source": source,
        "itemCount": len(items),
        "totalCost": round(sum(float(i.get("totalCost") or 0) for i in items), 2),
        "byGl": {k: round(v, 2) for k, v in sorted(by_gl.items())},
        "sheetsUsed": sheets_used,
        "sheetsSkipped": sheets_skipped,
        "items": items,
    }


def parse_bids_path(path: str | Path) -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return parse_bids_workbook(wb, source_name=path.name)
    finally:
        wb.close()


def parse_bids_bytes(data: bytes, filename: str = "bids.xlsx") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return parse_bids_workbook(wb, source_name=filename)
    finally:
        wb.close()
