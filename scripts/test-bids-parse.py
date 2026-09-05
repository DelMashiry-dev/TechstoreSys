"""Create a sample IT DIR-style bids workbook and verify bids_parse."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl
from bids_parse import parse_bids_path


def build_sample(path: Path) -> None:
    wb = openpyxl.Workbook()
    # toner sheet
    ws = wb.active
    ws.title = "toner 2026"
    ws.append(["IT DIR BIDS — Computer Consumables FY 2026"])
    ws.append([])
    ws.append(["Ser", "Item", "Description / Justification", "Cost Centre", "Qty", "Unit Cost", "Total Cost"])
    ws.append([1, "HP CE505A toner", "", "Z04P2SP212", 10, 130, 1300])
    ws.append([2, "HP Ink 652 black", "Inkjets", "Z04P2SP212", 5, 40, 200])
    ws.append([3, "TOTAL", "", "", "", "", 1500])

    ws2 = wb.create_sheet("ICT 2026")
    ws2.append(["Ser", "Item", "Cost Centre No.", "GL A/C", "Item Description", "Quantity Required", "Unit Cost", "Total Cost"])
    ws2.append([1, "Laptop HP EliteBook", "Z04P2SP212", "3112210001", "Duty staff issue", 2, 1500, 3000])
    ws2.append([2, "Desktop i5", "Z04P2SP212", "3112210001", "", 3, 800, 2400])

    ws3 = wb.create_sheet("SUMMARY")
    ws3.append(["Category", "Total"])
    ws3.append(["Toner", 1500])
    ws3.append(["ICT", 5400])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    sample = ROOT / "data" / "sample-it-dir-bids.xlsx"
    build_sample(sample)
    pack = parse_bids_path(sample)
    print("file", pack["file"])
    print("items", pack["itemCount"], "total", pack["totalCost"])
    print("sheetsUsed", pack["sheetsUsed"])
    print("byGl", pack["byGl"])
    assert pack["itemCount"] == 4, pack["itemCount"]
    assert abs(pack["totalCost"] - 6900) < 0.01, pack["totalCost"]
    print("OK", sample)


if __name__ == "__main__":
    main()
