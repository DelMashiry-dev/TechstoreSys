"""Parse IT DIR creditors Excel workbooks into supplier-debt case packs."""
from __future__ import annotations

import datetime
import io
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
}


def fmt_date(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v or "").strip()
    return s[:10] if s else ""


def num(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def slug(name: str) -> str:
    return "".join(ch if ch.isalnum() else "-" for ch in str(name or "").lower()).strip("-")


def parse_register_date(title: str, fallback: str = "") -> str:
    text = str(title or "")
    m = re.search(
        r"(\d{1,2})\s+(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+(\d{4})",
        text,
        re.I,
    )
    if not m:
        return fallback or datetime.date.today().isoformat()
    day, month_name, year = m.group(1), m.group(2).lower(), int(m.group(3))
    month = MONTHS.get(month_name, 11)
    try:
        return datetime.date(year, month, int(day)).isoformat()
    except ValueError:
        return fallback or datetime.date.today().isoformat()


def is_summary_row(company, vendor, po, inv, supply) -> bool:
    name = str(company or "").strip().lower()
    if not name or name == "`":
        if not str(vendor or "").strip():
            return True
        name = f"vendor {vendor}".lower()
    if name in {"grand total", "total", "sub total", "subtotal", "totals"}:
        return True
    if "grand total" in name:
        return True
    if not str(po or "").strip() and not str(inv or "").strip() and not fmt_date(supply):
        return True
    return False


def supplier_group_key(company: str, vendor: str) -> str:
    v = str(vendor or "").strip()
    if v and re.fullmatch(r"\d+", v):
        return f"vendor-{v}"
    return f"name-{slug(company)[:48]}" or "unknown"


def canonical_supplier_name(names: list[str]) -> str:
    cleaned = [str(n or "").strip() for n in names if str(n or "").strip()]
    if not cleaned:
        return "Unknown supplier"
    cleaned.sort(key=lambda n: (-len(n), n.lower()))
    return cleaned[0]


def read_register_title(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        if not row:
            continue
        for cell in row:
            text = str(cell or "").strip()
            if "CREDITOR" in text.upper():
                return text
    return ""


def parse_creditors_workbook(wb, *, source_name: str = "") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    ws = wb.active
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]

    title = read_register_title(ws)
    received_date = parse_register_date(title)
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    grouped: dict[str, dict] = {}
    for row in rows:
        if not row or not any(row):
            continue
        cells = list(row) + [None] * 15
        (
            _ser,
            supply,
            _cc,
            gl,
            company,
            vendor,
            po,
            inv,
            zwl,
            rate,
            conv,
            zwg,
            usd_inv,
            total_usd,
            remarks,
        ) = cells[:15]
        company = str(company or "").strip()
        if is_summary_row(company, vendor, po, inv, supply):
            continue
        if not company or company == "`":
            company = f"Vendor {vendor or 'unknown'}"
            if company.lower() == "vendor unknown":
                continue

        key = supplier_group_key(company, vendor)
        bucket = grouped.setdefault(key, {"names": [], "lines": []})
        bucket["names"].append(company)
        bucket["lines"].append(
            {
                "supplyDate": fmt_date(supply),
                "costCentre": "IT DIR",
                "poNo": str(po or "").strip(),
                "invoiceNo": str(inv or "").strip(),
                "amountZwl": num(zwl) or "",
                "rateUsd": str(rate or "").strip(),
                "convertedUsd": num(conv) or "",
                "amountZwg": num(zwg) or "",
                "amountUsd": num(usd_inv) or "",
                "totalUsd": round(num(total_usd), 2),
                "vendor": str(vendor or "").strip(),
                "gl": str(gl or "").strip(),
                "remarks": str(remarks or "").strip() if remarks else "",
            }
        )

    cases = []
    for key, bucket in sorted(grouped.items(), key=lambda item: item[0]):
        supplier = canonical_supplier_name(bucket["names"])
        lines = bucket["lines"]
        earliest = min((ln["supplyDate"] for ln in lines if ln["supplyDate"]), default="")
        total = round(sum(ln["totalUsd"] for ln in lines), 2)
        slug_part = slug(supplier)[:40] or key.replace("vendor-", "v-")
        cases.append(
            {
                "id": f"sd-it-cred-{slug_part}",
                "caseNo": f"IT-CR-{slug(supplier)[:28].upper()}",
                "supplier": supplier,
                "costCentre": "IT DIR",
                "receivedDate": received_date,
                "accumulatedFrom": earliest,
                "status": "open",
                "totalUsd": total,
                "lines": lines,
                "notes": f"Imported from IT DIR creditors register ({len(lines)} invoice line(s)).",
                "vendorKey": key,
            }
        )

    display_title = title.strip() or f"IT DIR CREDITORS AS AT {received_date}"
    return {
        "source": source_name or "creditors.xlsx",
        "asAt": received_date,
        "title": display_title,
        "caseCount": len(cases),
        "lineCount": sum(len(c["lines"]) for c in cases),
        "totalUsd": round(sum(c["totalUsd"] for c in cases), 2),
        "cases": cases,
    }


def parse_creditors_path(path: str | Path) -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return parse_creditors_workbook(wb, source_name=path.name)
    finally:
        wb.close()


def parse_creditors_bytes(data: bytes, filename: str = "creditors.xlsx") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return parse_creditors_workbook(wb, source_name=filename)
    finally:
        wb.close()
