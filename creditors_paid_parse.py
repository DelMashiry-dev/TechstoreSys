"""Parse DAF / finance paid-list Excel and match against creditor cases."""
from __future__ import annotations

import datetime
import io
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

COL_ALIASES = {
    "supplier": ("company", "supplier", "vendor", "creditor", "company name", "payee"),
    "po": ("po", "po number", "p/order", "purchase order", "order no", "order number"),
    "invoice": ("invoice", "invoice no", "inv", "invoice number"),
    "paid_date": ("paid", "paid date", "payment date", "date paid", "pay date", "settlement"),
    "amount": ("amount", "usd", "paid amount", "total usd", "total amount", "value"),
}


def fmt_date(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v or "").strip()
    return s[:10] if s else ""


def num(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def norm_header(cell) -> str:
    return re.sub(r"\s+", " ", str(cell or "").strip().lower())


def detect_paid_columns(headers: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        h = norm_header(raw)
        if not h:
            continue
        for key, aliases in COL_ALIASES.items():
            if key in mapping:
                continue
            if any(alias in h or h == alias for alias in aliases):
                mapping[key] = idx
    return mapping


def find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for i, row in enumerate(rows[:12]):
        headers = list(row or [])
        mapping = detect_paid_columns(headers)
        if mapping.get("supplier") is not None or mapping.get("invoice") is not None or mapping.get("po") is not None:
            if len(mapping) >= 2:
                return i, mapping
    return None, None


def parse_paid_workbook(wb, *, source_name: str = "") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx, mapping = find_header_row(rows)
    if header_idx is None or not mapping:
        raise ValueError("Could not find a paid-list header row (supplier / PO / invoice columns).")

    entries = []
    for row in rows[header_idx + 1 :]:
        if not row or not any(row):
            continue
        cells = list(row)

        def cell(key: str):
            idx = mapping.get(key)
            if idx is None or idx >= len(cells):
                return ""
            return cells[idx]

        supplier = str(cell("supplier") or "").strip()
        po = str(cell("po") or "").strip()
        invoice = str(cell("invoice") or "").strip()
        paid_date = fmt_date(cell("paid_date"))
        amount = round(num(cell("amount")), 2) if mapping.get("amount") is not None else 0.0

        if not supplier and not po and not invoice:
            continue
        if norm_header(supplier) in {"grand total", "total", "totals"}:
            continue

        entries.append(
            {
                "supplier": supplier,
                "poNo": po,
                "invoiceNo": invoice,
                "paidDate": paid_date,
                "amountUsd": amount,
            }
        )

    if not entries:
        raise ValueError("No paid rows found below the header.")

    return {
        "source": source_name or "paid-list.xlsx",
        "entryCount": len(entries),
        "entries": entries,
    }


def parse_paid_bytes(data: bytes, filename: str = "paid-list.xlsx") -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return parse_paid_workbook(wb, source_name=filename)
    finally:
        wb.close()


def parse_paid_path(path: str | Path) -> dict:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return parse_paid_workbook(wb, source_name=path.name)
    finally:
        wb.close()
